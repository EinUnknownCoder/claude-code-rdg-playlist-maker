"""Excel-Parsing Modul für RDG Playlist Maker."""

from openpyxl import load_workbook
from dataclasses import dataclass
import re


@dataclass
class Song:
    """Repräsentiert einen Song-Request aus der Excel-Datei."""
    youtube_url: str
    artist: str
    title: str
    description: str  # Songabschnitt, kann "Dancebreak" enthalten
    dancer_name: str
    start_seconds: float
    end_seconds: float
    row_number: int  # Zeilennummer in der Excel für Fehlermeldungen

    @property
    def is_dancebreak(self) -> bool:
        """Prüft ob dieser Song ein Dancebreak ist."""
        return "dancebreak" in self.description.lower()

    @property
    def filename(self) -> str:
        """Generiert einen sicheren Dateinamen für den Song."""
        # Entferne alle Sonderzeichen UND Leerzeichen, konvertiere zu lowercase
        safe_artist = re.sub(r'[^\w]', '', self.artist).lower()
        safe_title = re.sub(r'[^\w]', '', self.title).lower()
        return f"{safe_artist}-{safe_title}.mp3"

    def get_start_seconds(self) -> float:
        """Gibt Start-Zeit in Sekunden zurück."""
        return self.start_seconds

    def get_end_seconds(self) -> float:
        """Gibt End-Zeit in Sekunden zurück."""
        return self.end_seconds


def read_excel(filepath: str) -> list[Song]:
    """
    Liest die Song-Requests aus einer Excel-Datei.

    Erwartete Spalten (Reihenfolge):
    1. YouTube-URL
    2. Artist
    3. Title
    4. Description
    5. Requester/Dancer
    6. Start: Minute
    7. Start: Second
    8. End: Minute
    9. End: Seconds
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    songs = []

    # Überspringe Header-Zeile (Zeile 1)
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Prüfe ob Zeile leer ist
        if not row or not any(row):
            continue

        # Extrahiere Werte (mindestens 9 Spalten erwartet)
        values = list(row[:9])

        # Stelle sicher, dass wir genug Spalten haben
        while len(values) < 9:
            values.append(None)

        youtube_url = str(values[0]).strip() if values[0] else ""
        # Entferne Playlist-Parameter aus URL
        youtube_url = re.sub(r'&list=[^&]*', '', youtube_url)
        artist = str(values[1]).strip() if values[1] else ""
        title = str(values[2]).strip() if values[2] else ""
        description = str(values[3]).strip() if values[3] else ""
        dancer_name = str(values[4]).strip() if values[4] else ""

        # Timestamps: Minute und Sekunde getrennt
        try:
            start_min = int(values[5]) if values[5] is not None else 0
            start_sec = int(values[6]) if values[6] is not None else 0
            end_min = int(values[7]) if values[7] is not None else 0
            end_sec = int(values[8]) if values[8] is not None else 0
        except (ValueError, TypeError):
            print(f"Warnung: Ungültige Timestamps in Zeile {row_num}, überspringe...")
            continue

        start_seconds = start_min * 60 + start_sec
        end_seconds = end_min * 60 + end_sec

        # Überspringe Zeilen ohne URL
        if not youtube_url:
            continue

        song = Song(
            youtube_url=youtube_url,
            artist=artist,
            title=title,
            description=description,
            dancer_name=dancer_name,
            start_seconds=start_seconds,
            end_seconds=end_seconds,
            row_number=row_num
        )
        songs.append(song)

    wb.close()
    return songs


def validate_timestamps(songs: list[Song]) -> list[str]:
    """
    Validiert die Timestamps aller Songs.
    Gibt eine Liste von Fehlermeldungen zurück.
    """
    errors = []

    for song in songs:
        start = song.get_start_seconds()
        end = song.get_end_seconds()

        if start >= end:
            errors.append(
                f"Zeile {song.row_number}: Start-Zeit ({start}s) "
                f"muss vor End-Zeit ({end}s) liegen - "
                f"{song.artist} - {song.title}"
            )

    return errors
